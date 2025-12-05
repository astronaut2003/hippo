"""
实验数据汇总脚本

读取指定目录下所有 QA_X 子目录中的 score.json 和 retrieval.json，
汇总评分、token 占用、时间等指标，输出为 Excel 表格。
"""

import os
import json
import argparse
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def load_qa_results(qa_dir: Path) -> Dict[str, Any]:
    """
    加载单个 QA 目录的结果
    
    Args:
        qa_dir: QA 目录路径
        
    Returns:
        合并后的结果字典
    """
    result = {
        'qa_index': None,
        'status': 'missing'
    }
    
    # 从目录名提取索引
    dir_name = qa_dir.name
    if dir_name.startswith('QA_'):
        try:
            result['qa_index'] = int(dir_name[3:])
        except ValueError:
            result['qa_index'] = dir_name
    
    score_file = qa_dir / 'score.json'
    retrieval_file = qa_dir / 'retrieval.json'
    
    # 加载 score.json
    if score_file.exists():
        try:
            with open(score_file, 'r', encoding='utf-8') as f:
                score_data = json.load(f)
                result['score_data'] = score_data
                
                # ================= 修改开始 =================
                # 优先检查 evaluation_success。
                # 即使 status='failed' (例如因数据库只读导致)，只要评估成功完成，就视为 success。
                eval_info = score_data.get('evaluation', {})
                if isinstance(eval_info, dict) and eval_info.get('evaluation_success') is True:
                    result['status'] = 'success'
                else:
                    # 否则使用原始状态
                    result['status'] = score_data.get('status', 'unknown')
                # ================= 修改结束 =================
                
        except Exception as e:
            result['score_error'] = str(e)
            result['status'] = 'error'
    
    # 加载 retrieval.json
    if retrieval_file.exists():
        try:
            with open(retrieval_file, 'r', encoding='utf-8') as f:
                retrieval_data = json.load(f)
                result['retrieval_data'] = retrieval_data
        except Exception as e:
            result['retrieval_error'] = str(e)
    
    return result


def extract_metrics(result: Dict[str, Any]) -> Dict[str, Any]:
    """
    从结果中提取关键指标
    
    Args:
        result: load_qa_results 返回的结果
        
    Returns:
        扁平化的指标字典
    """
    metrics = {
        # 基本信息
        'qa_index': result.get('qa_index'),
        'status': result.get('status', 'unknown'),
        # 添加错误信息字段，方便排查
        'score_error': result.get('score_error', ''),
        'retrieval_error': result.get('retrieval_error', '')
    }
    
    score_data = result.get('score_data', {})
    retrieval_data = result.get('retrieval_data', {})
    
    # 问题信息
    metrics['question_id'] = score_data.get('question_id', '')
    metrics['question'] = score_data.get('question', '')
    metrics['question_type'] = score_data.get('question_type', '')
    metrics['gold_answer'] = score_data.get('gold_answer', '')
    metrics['predicted_answer'] = score_data.get('predicted_answer', '')
    metrics['question_date'] = score_data.get('question_date', '')
    
    # 模型信息
    metrics['gen_llm_model'] = score_data.get('gen_llm_model', '')
    metrics['eval_llm_model'] = score_data.get('eval_llm_model', '')
    
    # ===== 评分指标 =====
    scores = score_data.get('scores', {})
    
    # LLM 评分（主评分）
    metrics['llm_accuracy'] = scores.get('llm_accuracy', None)
    
    # 其他评分
    metrics['exact_match'] = scores.get('exact_match', None)
    metrics['token_f1'] = scores.get('token_f1', None)
    metrics['rouge1_f'] = scores.get('rouge1_f', None)
    metrics['rouge2_f'] = scores.get('rouge2_f', None)
    metrics['rougeL_f'] = scores.get('rougeL_f', None)
    metrics['semantic_similarity'] = scores.get('semantic_similarity', None)
    metrics['avg_lexical'] = scores.get('avg_lexical', None)
    metrics['avg_semantic'] = scores.get('avg_semantic', None)
    metrics['overall_average'] = scores.get('overall_average', None)
    
    # LLM 评估详情
    evaluation = score_data.get('evaluation', {})
    llm_details = evaluation.get('llm_details', {})
    metrics['llm_consistency'] = llm_details.get('consistency', None)
    metrics['llm_confidence'] = llm_details.get('confidence', '')
    
    # ===== Token 占用 =====
    gen_token = score_data.get('gen_token_usage', {})
    metrics['prompt_tokens'] = gen_token.get('prompt_tokens', None)
    metrics['answer_tokens'] = gen_token.get('answer_tokens', None)
    metrics['total_tokens'] = gen_token.get('total_tokens', None)
    metrics['context_length'] = gen_token.get('context_length', None)
    metrics['max_context_tokens'] = gen_token.get('max_context_tokens', None)
    metrics['prompt_ratio'] = gen_token.get('prompt_ratio', None)
    metrics['tokenizer_type'] = gen_token.get('tokenizer_type', '')
    metrics['encoding'] = gen_token.get('encoding', '')
    metrics['prompt_length_chars'] = score_data.get('prompt_length', None)
    
    # ===== 时间统计 =====
    timing = score_data.get('timing', {})
    metrics['load_time'] = timing.get('load_time', None)
    metrics['retrieval_time'] = timing.get('retrieval_time', None)
    metrics['generation_time'] = timing.get('generation_time', None)
    metrics['evaluation_time'] = timing.get('evaluation_time', None)
    metrics['cleanup_time'] = timing.get('cleanup_time', None)
    metrics['total_time'] = timing.get('total_time', None)
    
    # ===== 检索信息 =====
    metrics['query_top_k'] = retrieval_data.get('query_top_k', None)
    metrics['retrieved_memories_count'] = retrieval_data.get('retrieved_memories_count', None)
    
    load_result = retrieval_data.get('load_result', {})
    metrics['total_sessions'] = load_result.get('total_sessions', None)
    metrics['added_sessions'] = load_result.get('added_sessions', None)
    metrics['failed_sessions'] = load_result.get('failed_sessions', None)
    
    return metrics


def aggregate_results(results_dir: str, output_file: str = None) -> pd.DataFrame:
    """
    汇总目录下所有 QA 结果
    
    Args:
        results_dir: 结果目录路径
        output_file: 输出 Excel 文件路径（可选）
        
    Returns:
        汇总的 DataFrame
    """
    results_path = Path(results_dir)
    
    if not results_path.exists():
        raise FileNotFoundError(f"目录不存在: {results_dir}")
    
    # 查找所有 QA_X 目录
    qa_dirs = sorted(
        [d for d in results_path.iterdir() if d.is_dir() and d.name.startswith('QA_')],
        key=lambda x: int(x.name[3:]) if x.name[3:].isdigit() else 0
    )
    
    print(f"找到 {len(qa_dirs)} 个 QA 目录")
    
    # 收集所有指标
    all_metrics = []
    
    for qa_dir in qa_dirs:
        result = load_qa_results(qa_dir)
        metrics = extract_metrics(result)
        all_metrics.append(metrics)
    
    # 创建 DataFrame
    df = pd.DataFrame(all_metrics)
    
    # 按 qa_index 排序
    if 'qa_index' in df.columns:
        df = df.sort_values('qa_index').reset_index(drop=True)
    
    return df


def generate_summary_stats(df: pd.DataFrame) -> pd.DataFrame:
    """
    生成汇总统计
    
    Args:
        df: 原始数据 DataFrame
        
    Returns:
        统计摘要 DataFrame
    """
    stats = {}
    
    # 成功/失败统计
    stats['总样本数'] = len(df)
    stats['成功数'] = len(df[df['status'] == 'success'])
    stats['失败数'] = len(df[df['status'] != 'success'])
    stats['成功率'] = stats['成功数'] / stats['总样本数'] if stats['总样本数'] > 0 else 0
    
    # 仅对成功样本计算统计
    success_df = df[df['status'] == 'success']
    
    if len(success_df) > 0:
        # LLM 评分统计（主评分）
        if 'llm_accuracy' in success_df.columns:
            llm_scores = success_df['llm_accuracy'].dropna()
            stats['LLM准确率_平均'] = llm_scores.mean()
            stats['LLM准确率_中位数'] = llm_scores.median()
            stats['LLM准确率_标准差'] = llm_scores.std()
            stats['LLM正确数'] = (llm_scores == 1.0).sum()
            stats['LLM正确率'] = stats['LLM正确数'] / len(llm_scores) if len(llm_scores) > 0 else 0
        
        # 其他评分统计
        score_cols = ['exact_match', 'token_f1', 'rouge1_f', 'rougeL_f', 
                      'semantic_similarity', 'overall_average']
        for col in score_cols:
            if col in success_df.columns:
                values = success_df[col].dropna()
                if len(values) > 0:
                    stats[f'{col}_平均'] = values.mean()
        
        # Token 统计
        token_cols = ['prompt_tokens', 'answer_tokens', 'total_tokens']
        for col in token_cols:
            if col in success_df.columns:
                values = success_df[col].dropna()
                if len(values) > 0:
                    stats[f'{col}_平均'] = values.mean()
                    stats[f'{col}_总计'] = values.sum()
        
        # 时间统计
        time_cols = ['load_time', 'retrieval_time', 'generation_time', 
                     'evaluation_time', 'total_time']
        for col in time_cols:
            if col in success_df.columns:
                values = success_df[col].dropna()
                if len(values) > 0:
                    stats[f'{col}_平均(秒)'] = values.mean()
                    stats[f'{col}_总计(秒)'] = values.sum()
        
        # 检索统计
        if 'retrieved_memories_count' in success_df.columns:
            values = success_df['retrieved_memories_count'].dropna()
            if len(values) > 0:
                stats['检索记忆数_平均'] = values.mean()
    
    # 按问题类型统计 LLM 准确率
    if 'question_type' in success_df.columns and 'llm_accuracy' in success_df.columns:
        for q_type in success_df['question_type'].unique():
            if pd.notna(q_type):
                type_df = success_df[success_df['question_type'] == q_type]
                type_scores = type_df['llm_accuracy'].dropna()
                if len(type_scores) > 0:
                    stats[f'LLM准确率_{q_type}'] = type_scores.mean()
                    stats[f'样本数_{q_type}'] = len(type_scores)
    
    # 转换为 DataFrame
    stats_df = pd.DataFrame([stats]).T
    stats_df.columns = ['值']
    stats_df.index.name = '指标'
    
    return stats_df


def format_worksheet(ws, is_summary: bool = False):
    """
    格式化工作表
    
    Args:
        ws: openpyxl worksheet 对象
        is_summary: 是否为统计摘要表
    """
    # 定义样式
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    
    # 设置行高
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 22 if not is_summary else 28
    
    # 格式化表头（第一行）
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 格式化数据区域
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # 偶数行添加浅色背景
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color='E9EFF7', end_color='E9EFF7', fill_type='solid')
        else:
            row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for cell in row:
            cell.border = thin_border
            cell.fill = row_fill
            
            # 根据数据类型设置对齐
            if isinstance(cell.value, (int, float)):
                cell.alignment = number_alignment
                # 格式化数字
                if isinstance(cell.value, float):
                    if abs(cell.value) < 1 and cell.value != 0:
                        cell.number_format = '0.0000'
                    else:
                        cell.number_format = '#,##0.00'
            else:
                cell.alignment = cell_alignment
    
    # 自动调整列宽
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    # 计算显示长度（中文字符算2个宽度）
                    cell_str = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                    max_length = max(max_length, length)
            except:
                pass
        
        # 设置列宽，有最小和最大限制
        if is_summary:
            # 统计摘要表列宽更大
            adjusted_width = max(min(max_length + 4, 60), 25)
        else:
            adjusted_width = max(min(max_length + 2, 50), 10)
        
        ws.column_dimensions[column].width = adjusted_width


def format_summary_worksheet(ws):
    """
    专门格式化统计摘要工作表
    
    Args:
        ws: openpyxl worksheet 对象
    """
    # 定义样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    
    category_font = Font(bold=True, size=11, color='1F4E79')
    category_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    
    normal_font = Font(size=11)
    value_font = Font(size=11, bold=True, color='C00000')
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    value_alignment = Alignment(horizontal='right', vertical='center')
    
    thick_border = Border(
        left=Side(style='medium', color='2E75B6'),
        right=Side(style='medium', color='2E75B6'),
        top=Side(style='medium', color='2E75B6'),
        bottom=Side(style='medium', color='2E75B6')
    )
    
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 35  # 指标名称列
    ws.column_dimensions['B'].width = 25  # 值列
    
    # 定义分类关键词
    category_keywords = {
        '基本统计': ['总样本数', '成功数', '失败数', '成功率'],
        'LLM评分': ['LLM准确率', 'LLM正确数', 'LLM正确率'],
        '其他评分': ['exact_match', 'token_f1', 'rouge', 'semantic', 'overall'],
        'Token统计': ['tokens_平均', 'tokens_总计', 'prompt_tokens', 'answer_tokens', 'total_tokens'],
        '时间统计': ['time_平均', 'time_总计', 'load_time', 'retrieval_time', 'generation_time', 'evaluation_time', 'total_time'],
        '检索统计': ['检索记忆数'],
        '问题类型': ['样本数_']
    }
    
    # 格式化表头
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    ws.row_dimensions[1].height = 30
    
    # 格式化数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 26
        
        metric_name = str(row[0].value) if row[0].value else ''
        
        # 检查是否为分类标题行（通过关键词判断）
        is_category_start = False
        for cat_name, keywords in category_keywords.items():
            for keyword in keywords:
                if keyword in metric_name:
                    is_category_start = True
                    break
            if is_category_start:
                break
        
        for col_idx, cell in enumerate(row):
            cell.border = thin_border
            
            if col_idx == 0:  # 指标名称列
                cell.font = normal_font
                cell.alignment = cell_alignment
            else:  # 值列
                cell.alignment = value_alignment
                
                # 格式化数值
                if isinstance(cell.value, float):
                    if 'rate' in metric_name.lower() or '率' in metric_name or '准确' in metric_name:
                        # 百分比格式
                        cell.number_format = '0.00%'
                        cell.font = Font(size=11, bold=True, color='008000')  # 绿色
                    elif abs(cell.value) < 1 and cell.value != 0:
                        cell.number_format = '0.0000'
                        cell.font = value_font
                    else:
                        cell.number_format = '#,##0.00'
                        cell.font = value_font
                elif isinstance(cell.value, int):
                    cell.number_format = '#,##0'
                    cell.font = value_font
        
        # 偶数行添加浅色背景
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


def save_to_excel(
    df: pd.DataFrame, 
    stats_df: pd.DataFrame, 
    output_file: str,
    results_dir: str
):
    """
    保存结果到 Excel 文件（带格式化）
    
    Args:
        df: 详细数据 DataFrame
        stats_df: 统计摘要 DataFrame  
        output_file: 输出文件路径
        results_dir: 原始结果目录（用于记录元信息）
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: 统计摘要
        stats_df.to_excel(writer, sheet_name='统计摘要')
        
        # Sheet 2: 详细数据 - 核心指标
        core_cols = [
            'qa_index', 'status', 'question_id', 'question_type',
            'llm_accuracy', 'exact_match', 'token_f1', 'semantic_similarity',
            'prompt_tokens', 'answer_tokens', 'total_tokens',
            'load_time', 'retrieval_time', 'generation_time', 'total_time',
            'retrieved_memories_count', 'score_error', 'retrieval_error'
        ]
        core_df = df[[c for c in core_cols if c in df.columns]]
        core_df.to_excel(writer, sheet_name='核心指标', index=False)
        
        # Sheet 3: 详细数据 - 所有评分
        score_cols = [
            'qa_index', 'question_id', 'question_type',
            'llm_accuracy', 'llm_consistency', 'llm_confidence',
            'exact_match', 'token_f1', 
            'rouge1_f', 'rouge2_f', 'rougeL_f',
            'semantic_similarity', 'avg_lexical', 'avg_semantic', 'overall_average'
        ]
        score_df = df[[c for c in score_cols if c in df.columns]]
        score_df.to_excel(writer, sheet_name='评分详情', index=False)
        
        # Sheet 4: Token 和时间详情
        token_time_cols = [
            'qa_index', 'question_id',
            'prompt_tokens', 'answer_tokens', 'total_tokens',
            'prompt_length_chars', 'prompt_ratio',
            'load_time', 'retrieval_time', 'generation_time', 
            'evaluation_time', 'cleanup_time', 'total_time'
        ]
        token_time_df = df[[c for c in token_time_cols if c in df.columns]]
        token_time_df.to_excel(writer, sheet_name='Token与时间', index=False)
        
        # Sheet 5: 问答内容
        qa_cols = [
            'qa_index', 'question_id', 'question_type', 'question_date',
            'question', 'gold_answer', 'predicted_answer',
            'llm_accuracy', 'gen_llm_model', 'eval_llm_model'
        ]
        qa_df = df[[c for c in qa_cols if c in df.columns]]
        qa_df.to_excel(writer, sheet_name='问答内容', index=False)
        
        # Sheet 6: 完整数据
        df.to_excel(writer, sheet_name='完整数据', index=False)
        
        # Sheet 7: 元信息
        meta_info = {
            '结果目录': [results_dir],
            '生成时间': [datetime.now().isoformat()],
            'QA总数': [len(df)],
            '成功数': [len(df[df['status'] == 'success'])],
        }
        if 'gen_llm_model' in df.columns:
            meta_info['生成模型'] = [df['gen_llm_model'].iloc[0] if len(df) > 0 else '']
        if 'eval_llm_model' in df.columns:
            meta_info['评估模型'] = [df['eval_llm_model'].iloc[0] if len(df) > 0 else '']
        
        meta_df = pd.DataFrame(meta_info).T
        meta_df.columns = ['值']
        meta_df.index.name = '信息'
        meta_df.to_excel(writer, sheet_name='元信息')
        
        # 获取 workbook 并格式化各个工作表
        workbook = writer.book
        
        # 格式化统计摘要（特殊处理）
        format_summary_worksheet(workbook['统计摘要'])
        
        # 格式化其他工作表
        for sheet_name in ['核心指标', '评分详情', 'Token与时间', '问答内容', '完整数据', '元信息']:
            if sheet_name in workbook.sheetnames:
                format_worksheet(workbook[sheet_name], is_summary=(sheet_name == '元信息'))
    
    print(f"结果已保存到: {output_file}")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='汇总 LongMemEval Benchmark 实验数据，输出 Excel 表格'
    )
    parser.add_argument(
        '--input-dir', '-i',
        type=str,
        required=True,
        help='结果目录路径（包含 QA_X 子目录）'
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        default=None,
        help='输出 Excel 文件路径（默认保存到 benchmark_longmemeval/benchmark_results/）'
    )
    parser.add_argument(
        '--output-dir',
        type=str,
        default='benchmark_longmemeval/benchmark_results',
        help='输出目录（当 --output 未指定时使用）'
    )
    parser.add_argument(
        '--print-summary',
        action='store_true',
        help='打印统计摘要到终端'
    )
    
    args = parser.parse_args()
    
    # 汇总结果
    print(f"正在读取目录: {args.input_dir}")
    df = aggregate_results(args.input_dir)
    
    print(f"共读取 {len(df)} 条记录")
    print(f"成功: {len(df[df['status'] == 'success'])}, 失败: {len(df[df['status'] != 'success'])}")
    
    # 生成统计
    stats_df = generate_summary_stats(df)
    
    # 打印摘要
    if args.print_summary:
        print("\n" + "="*60)
        print("📊 统计摘要")
        print("="*60)
        print(stats_df.to_string())
        print("="*60)
    
    # ================= 打印失败样本详情 =================
    failed_df = df[df['status'] != 'success']
    if len(failed_df) > 0:
        print("\n" + "="*60)
        print(f"⚠️  发现 {len(failed_df)} 个失败样本 (Failed Samples)")
        print("="*60)
        # 按 qa_index 排序输出
        failed_df = failed_df.sort_values('qa_index')
        for _, row in failed_df.iterrows():
            qa_idx = row.get('qa_index', 'N/A')
            q_id = row.get('question_id') or 'Unknown'
            status = row.get('status', 'Unknown')
            
            print(f"Directory: QA_{qa_idx}")
            print(f"ID       : {q_id}")
            print(f"Status   : {status}")
            
            # 如果有具体的错误信息，则打印
            score_err = row.get('score_error')
            retr_err = row.get('retrieval_error')
            
            if score_err:
                print(f"Score Error: {score_err}")
            if retr_err:
                print(f"Retr Error : {retr_err}")
                
            print("-" * 30)
    # ===================================================

    # 生成输出文件名
    if args.output:
        output_file = args.output
    else:
        # 使用默认输出目录
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_dir_name = Path(args.input_dir).name
        output_file = str(output_dir / f"benchmark_summary_{input_dir_name}_{timestamp}.xlsx")
    
    # 保存到 Excel
    save_to_excel(df, stats_df, output_file, args.input_dir)
    
    # 打印关键指标
    success_df = df[df['status'] == 'success']
    if len(success_df) > 0 and 'llm_accuracy' in success_df.columns:
        llm_scores = success_df['llm_accuracy'].dropna()
        print(f"\n🎯 LLM 评估准确率: {llm_scores.mean():.2%} ({(llm_scores == 1.0).sum()}/{len(llm_scores)} 正确)")
    
    if len(success_df) > 0 and 'total_time' in success_df.columns:
        total_time = success_df['total_time'].sum()
        avg_time = success_df['total_time'].mean()
        print(f"⏱️  总耗时: {total_time:.2f}秒, 平均: {avg_time:.2f}秒/样本")
    
    if len(success_df) > 0 and 'total_tokens' in success_df.columns:
        total_tokens = success_df['total_tokens'].sum()
        print(f"📝 总 Token 数: {total_tokens:,.0f}")
    
    return df, stats_df


if __name__ == "__main__":
    main()